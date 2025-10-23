#!/usr/bin/env python3
"""
Excel Manager for JAIBird SENS Export
Manages creation and updating of Excel spreadsheets with SENS announcement data.
"""

import os
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from src.database.models import SensAnnouncement

logger = logging.getLogger(__name__)


class ExcelManagerError(Exception):
    """Custom exception for Excel Manager errors."""
    pass


class ExcelManager:
    """Manages Excel export of SENS announcements."""
    
    def __init__(self, excel_file_path: str = "data/sens_announcements.xlsx"):
        """
        Initialize Excel Manager.
        
        Args:
            excel_file_path: Path to the Excel file
        """
        self.excel_file_path = Path(excel_file_path)
        self.excel_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Excel styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.date_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        logger.info(f"Excel Manager initialized with file: {self.excel_file_path}")
    
    def create_or_update_spreadsheet(self, announcements: List[SensAnnouncement]) -> str:
        """
        Create or update Excel spreadsheet with SENS announcements.
        New announcements are added at the top.
        
        Args:
            announcements: List of SENS announcements to add/update
            
        Returns:
            str: Path to the created/updated Excel file
        """
        try:
            # Convert announcements to DataFrame
            df_new = self._announcements_to_dataframe(announcements)
            
            if df_new.empty:
                logger.info("No announcements to export")
                return str(self.excel_file_path)
            
            # Load existing data or create new
            if self.excel_file_path.exists():
                df_existing = self._load_existing_data()
                # Combine new and existing data (new on top)
                df_combined = self._merge_dataframes(df_new, df_existing)
            else:
                df_combined = df_new
            
            # Create Excel file
            self._create_excel_file(df_combined)
            
            logger.info(f"Excel file updated: {self.excel_file_path} ({len(df_combined)} total records)")
            return str(self.excel_file_path)
            
        except Exception as e:
            logger.error(f"Failed to create/update Excel spreadsheet: {e}")
            raise ExcelManagerError(f"Excel export failed: {e}")
    
    def _announcements_to_dataframe(self, announcements: List[SensAnnouncement]) -> pd.DataFrame:
        """Convert SENS announcements to pandas DataFrame."""
        if not announcements:
            return pd.DataFrame()
        
        data = []
        for announcement in announcements:
            # Determine PDF path (local if exists, otherwise online)
            pdf_path = self._get_pdf_path(announcement)
            
            data.append({
                'Date': announcement.date_published.strftime('%Y-%m-%d %H:%M') if announcement.date_published else '',
                'SENS Number': announcement.sens_number,
                'Organization': announcement.company_name,
                'Heading': announcement.title,
                'PDF Link': pdf_path,
                'PDF Summary': '',  # Placeholder for future AI summaries
                'Urgent': 'YES' if announcement.is_urgent else 'NO',
                'Created': announcement.date_scraped.strftime('%Y-%m-%d %H:%M:%S') if announcement.date_scraped else ''
            })
        
        df = pd.DataFrame(data)
        # Sort by date (newest first)
        if not df.empty and 'Date' in df.columns:
            df['Date_Sort'] = pd.to_datetime(df['Date'], errors='coerce')
            df = df.sort_values('Date_Sort', ascending=False, na_position='last')
            df = df.drop('Date_Sort', axis=1)
        
        return df
    
    def _get_pdf_path(self, announcement: SensAnnouncement) -> str:
        """Get the best available PDF path (local preferred, fallback to online)."""
        # Check for local PDF file
        local_pdf_patterns = [
            f"data/sens_pdfs/temp/SENS_{announcement.sens_number}*.pdf",
            f"data/sens_pdfs/{announcement.date_scraped.year if announcement.date_scraped else 2025}/{announcement.date_scraped.month if announcement.date_scraped else 9}/SENS_{announcement.sens_number}*.pdf"
        ]
        
        for pattern in local_pdf_patterns:
            from glob import glob
            matches = glob(pattern)
            if matches:
                # Convert to relative path for portability
                return os.path.relpath(matches[0])
        
        # Fallback to online URL if available
        if hasattr(announcement, 'pdf_url') and announcement.pdf_url:
            return announcement.pdf_url
        
        # Last resort: construct likely JSE URL
        return f"https://clientportal.jse.co.za/sens/{announcement.sens_number}"
    
    def _load_existing_data(self) -> pd.DataFrame:
        """Load existing Excel data."""
        try:
            df = pd.read_excel(self.excel_file_path)
            logger.debug(f"Loaded {len(df)} existing records from Excel")
            return df
        except Exception as e:
            logger.warning(f"Could not load existing Excel file: {e}")
            return pd.DataFrame()
    
    def _merge_dataframes(self, df_new: pd.DataFrame, df_existing: pd.DataFrame) -> pd.DataFrame:
        """Merge new and existing data, avoiding duplicates."""
        if df_existing.empty:
            return df_new
        
        if df_new.empty:
            return df_existing
        
        # Remove duplicates based on SENS Number
        df_existing_filtered = df_existing[~df_existing['SENS Number'].isin(df_new['SENS Number'])]
        
        # Combine: new on top
        df_combined = pd.concat([df_new, df_existing_filtered], ignore_index=True)
        
        logger.debug(f"Merged data: {len(df_new)} new + {len(df_existing_filtered)} existing = {len(df_combined)} total")
        return df_combined
    
    def _create_excel_file(self, df: pd.DataFrame):
        """Create formatted Excel file."""
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "SENS Announcements"
        
        # Add data to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Apply formatting
                if r_idx == 1:  # Header row
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = self.date_font
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                cell.border = self.border
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)
        
        # Add metadata sheet
        self._add_metadata_sheet(wb, len(df))
        
        # Save workbook
        wb.save(self.excel_file_path)
        logger.debug(f"Excel file saved with {len(df)} records")
    
    def _adjust_column_widths(self, ws: Worksheet):
        """Auto-adjust column widths for better readability."""
        column_widths = {
            'A': 18,  # Date
            'B': 12,  # SENS Number
            'C': 25,  # Organization
            'D': 50,  # Heading
            'E': 30,  # PDF Link
            'F': 40,  # PDF Summary (placeholder)
            'G': 8,   # Urgent
            'H': 18   # Created
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _add_metadata_sheet(self, wb: Workbook, record_count: int):
        """Add metadata sheet with export information."""
        meta_ws = wb.create_sheet("Export Info")
        
        metadata = [
            ['Export Information', ''],
            ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Records', record_count],
            ['Data Source', 'JAIBird SENS Scraper'],
            ['File Location', str(self.excel_file_path.absolute())],
            ['', ''],
            ['Column Descriptions', ''],
            ['Date', 'Publication date and time of SENS announcement'],
            ['SENS Number', 'Unique JSE SENS identifier (e.g., S510561)'],
            ['Organization', 'Company or entity that published the SENS'],
            ['Heading', 'Title/subject of the SENS announcement'],
            ['PDF Link', 'Path to PDF file (local preferred, online fallback)'],
            ['PDF Summary', 'AI-generated summary (future feature)'],
            ['Urgent', 'Whether announcement was flagged as urgent'],
            ['Created', 'When record was added to JAIBird database']
        ]
        
        for row_idx, (key, value) in enumerate(metadata, 1):
            meta_ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True) if value else Font()
            meta_ws.cell(row=row_idx, column=2, value=value)
        
        # Adjust metadata sheet column widths
        meta_ws.column_dimensions['A'].width = 20
        meta_ws.column_dimensions['B'].width = 40
    
    def export_filtered_data(self, filter_criteria: Dict[str, Any]) -> str:
        """
        Export filtered SENS data to Excel.
        
        Args:
            filter_criteria: Dictionary with filter parameters
                - company_name: Filter by company name
                - date_from: Start date filter
                - date_to: End date filter
                - is_urgent: Filter by urgent status
                - sens_numbers: List of specific SENS numbers
        
        Returns:
            str: Path to the exported Excel file
        """
        # This would integrate with database queries
        # For now, placeholder implementation
        logger.info(f"Filtered export requested with criteria: {filter_criteria}")
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filtered_file = self.excel_file_path.parent / f"sens_filtered_{timestamp}.xlsx"
        
        # Implementation would go here to query database and export
        logger.info(f"Filtered export would be saved to: {filtered_file}")
        return str(filtered_file)


def create_sens_excel_export(announcements: List[SensAnnouncement], 
                           excel_file_path: Optional[str] = None) -> str:
    """
    Convenience function to create SENS Excel export.
    
    Args:
        announcements: List of SENS announcements to export
        excel_file_path: Optional custom path for Excel file
    
    Returns:
        str: Path to the created Excel file
    """
    manager = ExcelManager(excel_file_path or "data/sens_announcements.xlsx")
    return manager.create_or_update_spreadsheet(announcements)


if __name__ == "__main__":
    # Test/demo code
    print("Excel Manager - SENS Export Tool")
    print("This module handles Excel export of SENS announcements")
